"""
===========================================
Reports App — Export Utilities
===========================================
PDF and Excel export for financial reports.
Uses ReportLab for PDFs and openpyxl for Excel.
"""

import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


CURRENCY_PREFIX = 'GH₵'


def export_report_to_pdf(report_data, report_type='balance_sheet'):
    """
    Export a financial report to PDF.

    Args:
        report_data: Dict from a report generator service
        report_type: Type of report for formatting

    Returns:
        HttpResponse with PDF content
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch)
    elements = []
    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,  # Center
        spaceAfter=20,
    )

    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,
        textColor=colors.grey,
        spaceAfter=30,
    )

    # Title
    elements.append(Paragraph(report_data.get('title', 'Financial Report'), title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        subtitle_style
    ))

    # Build table based on report type
    if report_type == 'balance_sheet':
        elements = _build_balance_sheet_pdf(elements, report_data, styles)
    elif report_type == 'income_statement':
        elements = _build_income_statement_pdf(elements, report_data, styles)
    elif report_type == 'cash_flow':
        elements = _build_cash_flow_pdf(elements, report_data, styles)
    elif report_type == 'budget_vs_actual':
        elements = _build_budget_vs_actual_pdf(elements, report_data, styles)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{datetime.now():%Y%m%d}.pdf"'
    return response


def _build_balance_sheet_pdf(elements, data, styles):
    """Build Balance Sheet PDF tables."""
    for section_name, section_key in [('ASSETS', 'assets'), ('LIABILITIES', 'liabilities'), ('EQUITY', 'equity')]:
        elements.append(Paragraph(section_name, styles['Heading2']))
        section = data.get(section_key, {})
        accounts = section.get('accounts', [])

        if accounts:
            table_data = [['Account Code', 'Account Name', 'Balance']]
            for acc in accounts:
                table_data.append([acc['code'], acc['name'], f"{CURRENCY_PREFIX}{acc['balance']:,.2f}"])
            table_data.append(['', f'Total {section_name}', f"{CURRENCY_PREFIX}{section['total']:,.2f}"])

            table = Table(table_data, colWidths=[1.5 * inch, 3.5 * inch, 1.5 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ]))
            elements.append(table)
        elements.append(Spacer(1, 15))

    return elements


def _build_income_statement_pdf(elements, data, styles):
    """Build Income Statement PDF."""
    for section_name, section_key in [('REVENUE', 'income'), ('EXPENSES', 'expenses')]:
        elements.append(Paragraph(section_name, styles['Heading2']))
        section = data.get(section_key, {})
        accounts = section.get('accounts', [])

        if accounts:
            table_data = [['Code', 'Account', 'Amount']]
            for acc in accounts:
                table_data.append([acc['code'], acc['name'], f"{CURRENCY_PREFIX}{acc['amount']:,.2f}"])
            table_data.append(['', f'Total {section_name}', f"{CURRENCY_PREFIX}{section['total']:,.2f}"])

            table = Table(table_data, colWidths=[1.2 * inch, 3.8 * inch, 1.5 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ]))
            elements.append(table)
        elements.append(Spacer(1, 15))

    # Net Income
    elements.append(Paragraph(
        f"<b>Net Income: {CURRENCY_PREFIX}{data.get('net_income', 0):,.2f}</b>",
        styles['Heading2']
    ))

    return elements


def _build_cash_flow_pdf(elements, data, styles):
    """Build Cash Flow PDF."""
    elements.append(Paragraph(f"Period: {data.get('start_date')} to {data.get('end_date')}", styles['Normal']))
    elements.append(Spacer(1, 10))

    # Summary table
    summary_data = [
        ['Cash Flow Summary', 'Amount'],
        ['Total Inflows', f"{CURRENCY_PREFIX}{data['inflows']['total']:,.2f}"],
        ['Total Outflows', f"{CURRENCY_PREFIX}{data['outflows']['total']:,.2f}"],
        ['Net Cash Flow', f"{CURRENCY_PREFIX}{data['net_cash_flow']:,.2f}"],
    ]
    table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(table)

    return elements


def _build_budget_vs_actual_pdf(elements, data, styles):
    """Build Budget vs Actual PDF."""
    table_data = [['Category', 'Budgeted', 'Actual', 'Variance', 'Utilization']]
    for item in data.get('items', []):
        table_data.append([
            item['category'],
            f"{CURRENCY_PREFIX}{item['budgeted']:,.2f}",
            f"{CURRENCY_PREFIX}{item['actual']:,.2f}",
            f"{CURRENCY_PREFIX}{item['variance']:,.2f}",
            f"{item['utilization']}%",
        ])

    if data.get('summary'):
        s = data['summary']
        table_data.append([
            'TOTAL',
            f"{CURRENCY_PREFIX}{s['total_budgeted']:,.2f}",
            f"{CURRENCY_PREFIX}{s['total_actual']:,.2f}",
            f"{CURRENCY_PREFIX}{s['total_variance']:,.2f}",
            f"{s['overall_utilization']}%",
        ])

    table = Table(table_data, colWidths=[2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(table)

    return elements


def export_report_to_excel(report_data, report_type='balance_sheet'):
    """
    Export a financial report to Excel (.xlsx).

    Args:
        report_data: Dict from a report generator service
        report_type: Type of report

    Returns:
        HttpResponse with Excel content
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_data.get('title', 'Report')

    # Styles
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='111827', end_color='111827', fill_type='solid')
    subheader_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    subheader_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    total_font = Font(name='Calibri', bold=True)
    total_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    currency_format = '"GH₵"#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Title row
    ws.merge_cells('A1:E1')
    ws['A1'] = report_data.get('title', 'Financial Report')
    ws['A1'].font = Font(name='Calibri', size=16, bold=True)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=9, color='888888')

    row = 4

    if report_type == 'balance_sheet':
        for section_name, section_key in [('ASSETS', 'assets'), ('LIABILITIES', 'liabilities'), ('EQUITY', 'equity')]:
            ws.cell(row=row, column=1, value=section_name).font = Font(bold=True, size=12)
            row += 1

            # Sub-header
            for col, header_text in enumerate(['Code', 'Account Name', 'Balance'], 1):
                cell = ws.cell(row=row, column=col, value=header_text)
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.border = thin_border
            row += 1

            section = report_data.get(section_key, {})
            for acc in section.get('accounts', []):
                ws.cell(row=row, column=1, value=acc['code']).border = thin_border
                ws.cell(row=row, column=2, value=acc['name']).border = thin_border
                cell = ws.cell(row=row, column=3, value=acc['balance'])
                cell.number_format = currency_format
                cell.border = thin_border
                row += 1

            # Total
            ws.cell(row=row, column=2, value=f'Total {section_name}').font = total_font
            ws.cell(row=row, column=2).fill = total_fill
            cell = ws.cell(row=row, column=3, value=section.get('total', 0))
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = currency_format
            row += 2

    elif report_type == 'income_statement':
        for section_name, section_key in [('REVENUE', 'income'), ('EXPENSES', 'expenses')]:
            ws.cell(row=row, column=1, value=section_name).font = Font(bold=True, size=12)
            row += 1
            for col, header_text in enumerate(['Code', 'Account', 'Amount'], 1):
                cell = ws.cell(row=row, column=col, value=header_text)
                cell.font = subheader_font
                cell.fill = subheader_fill
            row += 1
            section = report_data.get(section_key, {})
            for acc in section.get('accounts', []):
                ws.cell(row=row, column=1, value=acc['code'])
                ws.cell(row=row, column=2, value=acc['name'])
                ws.cell(row=row, column=3, value=acc['amount']).number_format = currency_format
                row += 1
            ws.cell(row=row, column=2, value=f'Total {section_name}').font = total_font
            ws.cell(row=row, column=3, value=section.get('total', 0)).number_format = currency_format
            row += 2

        ws.cell(row=row, column=2, value='NET INCOME').font = Font(bold=True, size=12)
        ws.cell(row=row, column=3, value=report_data.get('net_income', 0)).number_format = currency_format

    elif report_type == 'cash_flow':
        # Summary section
        ws.cell(row=row, column=1, value='CASH FLOW SUMMARY').font = Font(bold=True, size=12)
        row += 1

        for col, header_text in enumerate(['Description', 'Amount'], 1):
            cell = ws.cell(row=row, column=col, value=header_text)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1

        summary_rows = [
            ('Total Inflows', report_data.get('inflows', {}).get('total', 0)),
            ('Total Outflows', report_data.get('outflows', {}).get('total', 0)),
            ('Net Cash Flow', report_data.get('net_cash_flow', 0)),
        ]
        for label, amount in summary_rows:
            ws.cell(row=row, column=1, value=label).border = thin_border
            cell = ws.cell(row=row, column=2, value=amount)
            cell.number_format = currency_format
            cell.border = thin_border
            row += 1

        # Net Cash Flow total row styling
        ws.cell(row=row - 1, column=1).font = total_font
        ws.cell(row=row - 1, column=1).fill = total_fill
        ws.cell(row=row - 1, column=2).font = total_font
        ws.cell(row=row - 1, column=2).fill = total_fill
        row += 1

        # Inflows breakdown
        ws.cell(row=row, column=1, value='INFLOWS BY PAYMENT METHOD').font = Font(bold=True, size=11)
        row += 1
        for col, header_text in enumerate(['Payment Method', 'Amount'], 1):
            cell = ws.cell(row=row, column=col, value=header_text)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1
        for item in report_data.get('inflows', {}).get('breakdown', []):
            ws.cell(row=row, column=1, value=str(item.get('method', '')).title()).border = thin_border
            cell = ws.cell(row=row, column=2, value=item.get('amount', 0))
            cell.number_format = currency_format
            cell.border = thin_border
            row += 1
        row += 1

        # Outflows breakdown
        ws.cell(row=row, column=1, value='OUTFLOWS BY CATEGORY').font = Font(bold=True, size=11)
        row += 1
        for col, header_text in enumerate(['Category', 'Amount'], 1):
            cell = ws.cell(row=row, column=col, value=header_text)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1
        for item in report_data.get('outflows', {}).get('breakdown', []):
            ws.cell(row=row, column=1, value=item.get('category', '')).border = thin_border
            cell = ws.cell(row=row, column=2, value=item.get('amount', 0))
            cell.number_format = currency_format
            cell.border = thin_border
            row += 1

    elif report_type == 'budget_vs_actual':
        headers = ['Category', 'Budgeted', 'Actual', 'Variance', 'Utilization %']
        for col, header_text in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header_text)
            cell.font = subheader_font
            cell.fill = subheader_fill
        row += 1
        for item in report_data.get('items', []):
            ws.cell(row=row, column=1, value=item['category'])
            ws.cell(row=row, column=2, value=item['budgeted']).number_format = currency_format
            ws.cell(row=row, column=3, value=item['actual']).number_format = currency_format
            ws.cell(row=row, column=4, value=item['variance']).number_format = currency_format
            ws.cell(row=row, column=5, value=item['utilization'])
            row += 1

    # Auto-fit columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{datetime.now():%Y%m%d}.xlsx"'
    return response
